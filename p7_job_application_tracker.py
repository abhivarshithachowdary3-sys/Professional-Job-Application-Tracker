"""
Job Application Tracker
========================
Track job applications, interviews, and follow-ups with statistics and reminders.

Author: Abhivarshitha Chowdary
Date: February 2026
Python Version: 3.x
"""

import json
import os
from datetime import datetime, timedelta

# File to store applications
APPLICATIONS_FILE = "job_applications.json"

# Status options
STATUSES = [
    "applied",
    "under_review",
    "interview_scheduled",
    "interviewed",
    "offer",
    "rejected",
    "withdrawn",
    "ghosted"
]

# Status display with emojis
STATUS_DISPLAY = {
    "applied": "📝 Applied",
    "under_review": "👀 Under Review",
    "interview_scheduled": "📅 Interview Scheduled",
    "interviewed": "✅ Interviewed",
    "offer": "🎉 Offer Received",
    "rejected": "❌ Rejected",
    "withdrawn": "🚫 Withdrawn",
    "ghosted": "👻 Ghosted"
}

def load_applications():
    """Load applications from JSON file"""
    if os.path.exists(APPLICATIONS_FILE):
        with open(APPLICATIONS_FILE, "r") as file:
            return json.load(file)
    return []

def save_applications(applications):
    """Save applications to JSON file"""
    with open(APPLICATIONS_FILE, "w") as file:
        json.dump(applications, file, indent=4)

def generate_id(applications):
    """Generate unique ID for new application"""
    if len(applications) == 0:
        return 1
    return max(app["id"] for app in applications) + 1

# Load existing applications
applications = load_applications()

print("="*60)
print("          JOB APPLICATION TRACKER")
print("="*60)

def add_application():
    """Add a new job application"""
    print("\n--- ADD NEW APPLICATION ---")
    
    # Company name
    company = input("Company name: ").strip()
    if company == "":
        print("Company name cannot be empty!")
        return
    
    # Position
    position = input("Position/Role: ").strip()
    if position == "":
        print("Position cannot be empty!")
        return
    
    # Location (optional)
    location = input("Location (optional): ").strip()
    
    # Date applied
    date_input = input("Date applied (press Enter for today, or YYYY-MM-DD): ").strip()
    if date_input == "":
        date_applied = datetime.now().strftime("%Y-%m-%d")
    else:
        date_applied = date_input
    
    # Salary range (optional)
    salary = input("Salary range (e.g., 80k-100k, optional): ").strip()
    
    # Job URL (optional)
    job_url = input("Job posting URL (optional): ").strip()
    
    # Notes (optional)
    notes = input("Notes (optional): ").strip()
    
    # Follow-up date
    follow_up_input = input("Follow-up date (optional, YYYY-MM-DD): ").strip()
    follow_up_date = follow_up_input if follow_up_input else None
    
    # Create application
    application = {
        "id": generate_id(applications),
        "company": company,
        "position": position,
        "location": location,
        "date_applied": date_applied,
        "status": "applied",  # Default status
        "salary_range": salary,
        "job_url": job_url,
        "notes": notes,
        "follow_up_date": follow_up_date,
        "interview_date": None,
        "contact_person": None,
        "contact_email": None,
        "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    applications.append(application)
    save_applications(applications)
    
    print(f"\n✓ Application to {company} for {position} added!")
    print(f"  Status: {STATUS_DISPLAY['applied']}")

def view_all_applications():
    """Display all applications sorted by date"""
    if len(applications) == 0:
        print("\nNo applications yet! Add your first application.")
        return
    
    print("\n" + "="*80)
    print("                       ALL APPLICATIONS")
    print("="*80)
    
    # Sort by date (newest first)
    sorted_apps = sorted(applications, key=lambda x: x["date_applied"], reverse=True)
    
    for app in sorted_apps:
        status_icon = STATUS_DISPLAY[app["status"]]
        
        print(f"\n{status_icon} | ID: {app['id']} | Applied: {app['date_applied']}")
        print(f"   📍 {app['company']} - {app['position']}")
        
        if app['location']:
            print(f"   🌍 Location: {app['location']}")
        
        if app['salary_range']:
            print(f"   💰 Salary: {app['salary_range']}")
        
        if app['follow_up_date']:
            print(f"   🔔 Follow-up: {app['follow_up_date']}")
        
        if app['interview_date']:
            print(f"   📅 Interview: {app['interview_date']}")
        
        if app['notes']:
            print(f"   📝 Notes: {app['notes']}")
    
    print("\n" + "="*80)

def view_by_status():
    """Display applications grouped by status"""
    if len(applications) == 0:
        print("\nNo applications yet!")
        return
    
    print("\n" + "="*80)
    print("                   APPLICATIONS BY STATUS")
    print("="*80)
    
    # Group by status
    by_status = {}
    for status in STATUSES:
        by_status[status] = []
    
    for app in applications:
        by_status[app["status"]].append(app)
    
    # Display each status group
    for status in STATUSES:
        apps = by_status[status]
        if len(apps) > 0:
            print(f"\n{STATUS_DISPLAY[status]} ({len(apps)})")
            print("-" * 80)
            
            for app in sorted(apps, key=lambda x: x["date_applied"], reverse=True):
                print(f"  • {app['company']} - {app['position']} (Applied: {app['date_applied']})")
    
    print("\n" + "="*80)

def update_status():
    """Update the status of an application"""
    if len(applications) == 0:
        print("\nNo applications to update!")
        return
    
    # Show all applications
    print("\n--- YOUR APPLICATIONS ---")
    for app in applications:
        print(f"ID {app['id']}: {app['company']} - {app['position']} [{STATUS_DISPLAY[app['status']]}]")
    
    # Get application ID
    try:
        app_id = int(input("\nEnter application ID to update: "))
        
        # Find application
        app = None
        for a in applications:
            if a["id"] == app_id:
                app = a
                break
        
        if not app:
            print("Application not found!")
            return
        
        # Show current status
        print(f"\nCurrent status: {STATUS_DISPLAY[app['status']]}")
        
        # Show status options
        print("\nNew status:")
        for i, status in enumerate(STATUSES, 1):
            print(f"{i}. {STATUS_DISPLAY[status]}")
        
        # Get new status
        choice = int(input("\nChoose new status (1-8): "))
        
        if 1 <= choice <= len(STATUSES):
            new_status = STATUSES[choice - 1]
            app["status"] = new_status
            app["last_updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # If interview scheduled, ask for date
            if new_status == "interview_scheduled":
                interview_date = input("Interview date (YYYY-MM-DD): ").strip()
                if interview_date:
                    app["interview_date"] = interview_date
            
            save_applications(applications)
            print(f"\n✓ Status updated to: {STATUS_DISPLAY[new_status]}")
        else:
            print("Invalid choice!")
    
    except ValueError:
        print("Please enter a valid number!")

def dashboard():
    """Show quick dashboard with key metrics"""
    if len(applications) == 0:
        print("\nNo data yet! Start tracking applications.")
        return
    
    print("\n" + "="*80)
    print("                         DASHBOARD")
    print("="*80)
    
    # Calculate metrics
    total = len(applications)
    by_status = {}
    for status in STATUSES:
        by_status[status] = len([a for a in applications if a["status"] == status])
    
    # Active applications (not rejected/withdrawn/ghosted)
    active = total - by_status["rejected"] - by_status["withdrawn"] - by_status["ghosted"]
    
    # Response rate (got beyond "applied")
    responded = total - by_status["applied"] - by_status["ghosted"]
    response_rate = (responded / total * 100) if total > 0 else 0
    
    # Success rate (offers received)
    success_rate = (by_status["offer"] / total * 100) if total > 0 else 0
    
    print(f"\n📊 OVERVIEW")
    print(f"   Total Applications: {total}")
    print(f"   Active: {active}")
    print(f"   Response Rate: {response_rate:.1f}%")
    print(f"   Success Rate: {success_rate:.1f}%")
    
    print(f"\n📈 STATUS BREAKDOWN")
    for status in STATUSES:
        count = by_status[status]
        if count > 0:
            print(f"   {STATUS_DISPLAY[status]}: {count}")
    
    # Follow-ups needed today
    today = datetime.now().strftime("%Y-%m-%d")
    follow_ups_today = [a for a in applications if a.get("follow_up_date") == today]
    
    if follow_ups_today:
        print(f"\n🔔 FOLLOW-UPS TODAY ({len(follow_ups_today)})")
        for app in follow_ups_today:
            print(f"   • {app['company']} - {app['position']}")
    
    # Upcoming interviews
    upcoming_interviews = [a for a in applications if a.get("interview_date") and a["interview_date"] >= today]
    
    if upcoming_interviews:
        print(f"\n📅 UPCOMING INTERVIEWS ({len(upcoming_interviews)})")
        for app in sorted(upcoming_interviews, key=lambda x: x["interview_date"]):
            print(f"   • {app['interview_date']}: {app['company']} - {app['position']}")
    
    print("\n" + "="*80)

def delete_application():
    """Delete an application"""
    if len(applications) == 0:
        print("\nNo applications to delete!")
        return
    
    # Show all applications
    print("\n--- YOUR APPLICATIONS ---")
    for app in applications:
        print(f"ID {app['id']}: {app['company']} - {app['position']} [{STATUS_DISPLAY[app['status']]}]")
    
    # Get application ID
    try:
        app_id = int(input("\nEnter application ID to delete: "))
        
        # Find and remove application
        for i, app in enumerate(applications):
            if app["id"] == app_id:
                confirm = input(f"Delete {app['company']} - {app['position']}? (y/n): ").lower()
                
                if confirm == "y":
                    deleted = applications.pop(i)
                    save_applications(applications)
                    print(f"\n✓ Application to {deleted['company']} deleted!")
                else:
                    print("Deletion cancelled.")
                return
        
        print("Application not found!")
    
    except ValueError:
        print("Please enter a valid number!")


def export_to_excel():
    """Export all applications to Excel file"""
    if len(applications) == 0:
        print("\nNo applications to export!")
        return
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Applications"
        
        # Headers
        headers = [
            "ID", "Company", "Position", "Location", "Date Applied", 
            "Status", "Salary Range", "Interview Date", "Follow-up Date", 
            "Notes", "Job URL"
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row, app in enumerate(sorted(applications, key=lambda x: x["date_applied"], reverse=True), 2):
            ws.cell(row=row, column=1, value=app["id"])
            ws.cell(row=row, column=2, value=app["company"])
            ws.cell(row=row, column=3, value=app["position"])
            ws.cell(row=row, column=4, value=app["location"] or "")
            ws.cell(row=row, column=5, value=app["date_applied"])
            ws.cell(row=row, column=6, value=STATUS_DISPLAY[app["status"]])
            ws.cell(row=row, column=7, value=app["salary_range"] or "")
            ws.cell(row=row, column=8, value=app.get("interview_date") or "")
            ws.cell(row=row, column=9, value=app.get("follow_up_date") or "")
            ws.cell(row=row, column=10, value=app["notes"] or "")
            ws.cell(row=row, column=11, value=app["job_url"] or "")
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Save file
        filename = f"job_applications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        print(f"\n✓ Excel file exported successfully!")
        print(f"  Filename: {filename}")
        print(f"  Total applications: {len(applications)}")
        
    except ImportError:
        print("\n❌ openpyxl library not installed!")
        print("Install with: pip install openpyxl --break-system-packages")
    except Exception as e:
        print(f"\n❌ Error exporting to Excel: {e}")

# Load existing applications
applications = load_applications()


while True:
    print("\n--- MAIN MENU ---")
    print("1. 📊 Dashboard")
    print("2. ➕ Add Application")
    print("3. 📋 View All Applications")
    print("4. 🗂️  View by Status")
    print("5. ✏️  Update Application Status")
    print("6. 🗑️  Delete Application")
    print("7. 📤 Export to Excel")          # ← NEW
    print("8. 💾 Exit")                     # ← Changed from 7 to 8
    
    choice = input("\nChoose option (1-8): ")
    
    if choice == "1":
        dashboard()
    
    elif choice == "2":
        add_application()
    
    elif choice == "3":
        view_all_applications()
    
    elif choice == "4":
        view_by_status()
    
    elif choice == "5":
        update_status()
    
    elif choice == "6":
        delete_application()
    
    elif choice == "7":
        export_to_excel()
    
    elif choice == "8":
        print("\n💾 All applications saved!")
        print("Good luck with your job search! 🚀")
        break
    
    else:
        print("Invalid choice! Please enter 1-8.")